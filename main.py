from flask import Flask, request
from openpyxl import load_workbook
import openpyxl.styles
import datetime
from flask import send_file
import os

app = Flask(__name__)

TEMPLATE_PATH = "attached_assets/جدول_افلاین_عمر_و_تشکیل_سرمایه_رشد_1 (1)_1753971361572.xlsx"





USERNAME = "2011PHH90-bit"
PASSWORD = "@X9v#Lr8$zQ!mT2&fW7^bK#uE6@pY3$"


from flask import Flask, request, redirect, url_for, session
app.secret_key = "your_super_secret_key"
  # برای session

@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        if username == USERNAME and password == PASSWORD:
            session["logged_in"] = True
            return redirect(url_for("form"))
        else:
            error = "❌ نام کاربری یا رمز عبور اشتباه است."

    return f"""
    <!DOCTYPE html>
    <html lang="fa" dir="rtl">
    <head>
        <meta charset="UTF-8">
        <title>ورود به سامانه</title>
        <style>
            body {{
                font-family: 'Tahoma';
                background-color: #6CABDD;
                display: flex;
                justify-content: center;
                align-items: center;
                height: 100vh;
                direction: rtl;
                text-align: right;
            }}
            .login-box {{
                background-color: white;
                padding: 30px;
                border-radius: 10px;
                box-shadow: 0 0 10px rgba(0,0,0,0.3);
                width: 300px;
            }}
            h2 {{
                text-align: center;
                margin-bottom: 20px;
                color: #003366;
            }}
            label {{
                font-weight: bold;
                display: block;
                margin-top: 10px;
            }}
            input {{
                width: 100%;
                padding: 8px;
                margin-top: 5px;
                border: 1px solid #ccc;
                border-radius: 5px;
            }}
            button {{
                width: 100%;
                margin-top: 20px;
                padding: 10px;
                background-color: #003366;
                color: white;
                border: none;
                font-size: 16px;
                cursor: pointer;
                border-radius: 5px;
            }}
            .error {{
                color: red;
                margin-top: 10px;
                text-align: center;
            }}
        </style>
    </head>
    <body>
        <div class="login-box">
            <h2>ورود به سامانه بیمه سامان</h2>
            <form method="post">
                <label>نام کاربری:</label>
                <input type="text" name="username" required>
                <label>رمز عبور:</label>
                <input type="password" name="password" required>
                <button type="submit">ورود</button>
            </form>
            <div class="error">{error}</div>
        </div>
    </body>
    </html>
    """





dropdown_fields = {
    "روش پرداخت حق بیمه": ["ماهانه", "دوماهه", "سه ماهه", "چهارماهه", "شش ماهه", "سالانه"],
    "طرح": ["پرریسک", "رشد1", "رشد", "کم_ریسک", "متوسط_ریسک"],
    "ضریب درآمد ازکارافتادگی": ["1", "2", "3"],
    "منظور کردن پوشش ها در حق بیمه اولیه": ["بله", "خیر"],
    "دارای درامد از کارافتادگی": ["بله", "خیر"],
    "طرح امراض": ["معمولی", "پایه", "آسایش", "ممتاز"],
    "طبقه شغلی بیمه‌گذار": ["طبقه چهار", "طبقه دو از کارافتادگی", "طبقه دو کلی", "طبقه سه از کارافتادگی", "طبقه سه کلی", "طبقه یک", "طبقه پنج"],
    "دارای پوشش معافیت": ["بله", "خیر"],
    "طبقه شغلی بیمه‌شده": ["طبقه چهار", "طبقه دو از کارافتادگی", "طبقه دو کلی", "طبقه سه از کارافتادگی", "طبقه سه کلی", "طبقه یک", "طبقه پنج"]
}

interest_by_plan = {
    "پرریسک": "20.73%",
    "رشد1": "33.69%",
    "رشد": "22.69%",
    "کم_ریسک": "16.52%",
    "متوسط_ریسک": "17.18%"
}

mapping = {
    "سن بیمه‌گذار": "C6",
    "سن بیمه شده": "F6",
    "مدت بیمه نامه": "I6",
    "حق بیمه اولیه": "I8",
    "سرمایه اولیه": "I10",
    "درصد ضریب تغییر حق بيمه": "F8",
    "درصد ضریب تغییر سرمايه": "F10",
    "میانگین نرخ سود قطعی": "I22",
    "ضریب فوت حادثه": "I12",
    "ضریب هزینه پزشکی ناشی از حادثه": "C12",
    "ضریب درآمد ازکارافتادگی": "C14",
    "ضریب از کار افتادگی ناشی از حادثه": "F12",
    "مبلغ پرداخت اضافه": "C20",
    "شماره اولین سال در پرداخت اضافه": "I20",
    "شماره آخرین سال در پرداخت اضافه": "F20",
    "طرح": "C10",
    "روش پرداخت حق بیمه": "C8",
    "طرح امراض": "F16",
    "دارای پوشش معافیت": "I14",
    "دارای درامد از کارافتادگی": "F14",
    "منظور کردن پوشش ها در حق بیمه اولیه": "C16",
    "درصد اضافه نرخ پزشکی": "C18",
    "سرمایه امراض": "I16",
    "طبقه شغلی بیمه‌گذار": "F18",
    "طبقه شغلی بیمه‌شده": "I18"
}

def generate_filename():
    now = datetime.datetime.now()
    return f"جدول_خروجی_{now.strftime('%Y-%m-%d_%H-%M')}.xlsx"

def colorize_table(path):
    try:
        wb = load_workbook(path)
        sheet = wb["جدول نمونه"]
        sheet["A1"].fill = openpyxl.styles.PatternFill(start_color="00B4FF", end_color="00B4FF", fill_type="solid")
        sheet["A1"].font = openpyxl.styles.Font(color="FFFFFF", bold=True)

        dark_cells = [
            "A3", "B3", "C3", "C2", "D3", "D2", "E3", "Z3",
            "AA3", "AB3", "AB2", "AC3", "AC2", "AD3", "AE3",
            "AF3", "AH3", "AJ3"
        ]
        for cell in dark_cells:
            sheet[cell].fill = openpyxl.styles.PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            sheet[cell].font = openpyxl.styles.Font(color="FFFFFF")

        wb.save(path)
        wb.close()
    except Exception as e:
        print("❌ خطا در رنگ‌گذاری:", e)

def add_titles(path):
    try:
        wb = load_workbook(path)
        sheet = wb["جدول نمونه"]
        sheet["A1"].value = "جدول خروجی بیمه سامان"
        sheet["A1"].font = openpyxl.styles.Font(size=16, bold=True, color="FFFFFF")
        sheet["A100"].value = "Made By: POUYAN HEIDARI HERIS"
        sheet["A100"].font = openpyxl.styles.Font(size=6, color="666666")
        wb.save(path)
        wb.close()
    except Exception as e:
        print("❌ خطا در درج عنوان:", e)

@app.route("/", methods=["GET", "POST"])
def form():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    values = {field: "" for field in mapping.keys()}
    message = ""

    if request.method == "POST":
        action = request.form.get("action")
        values = {field: request.form.get(field, "") for field in mapping.keys()}

        if action == "ذخیره فرم":
            try:
                filename = generate_filename()
                wb = load_workbook(TEMPLATE_PATH)
                sheet = wb["Info"]
                for field, cell in mapping.items():
                    value = values[field]
                    try:
                        if value.replace('.', '', 1).isdigit():
                            sheet[cell] = float(value)
                        else:
                            sheet[cell] = value
                    except:
                        sheet[cell] = value
                wb.save(filename)
                colorize_table(filename)
                add_titles(filename)
                session["locked"] = True
                session["filename"] = filename
                message = "✅ فرم با موفقیت ذخیره شد. حالا می‌تونی فایل رو دانلود کنی."
            except Exception as e:
                message = f"❌ خطا در ذخیره‌سازی: {e}"

        elif action == "ویرایش فرم":
            session["locked"] = False
            session.pop("filename", None)
            message = "✅ فرم قابل ویرایش شد."

    locked = session.get("locked", False)
    return render_form_html(values, locked, message)

@app.route("/download")
def download():
    filename = session.get("filename")
    if filename and os.path.exists(filename):
        return send_file(filename, as_attachment=True)
    return "❌ فایل خروجی پیدا نشد."

def render_form_html(values, locked, message):
    form_html = ""
    count = 0
    row_html = ""

    for field in mapping.keys():
        label = f"<label>{field}</label>"
        disabled_attr = "disabled" if locked else ""

        if field in dropdown_fields:
            options = "".join([
                f'<option value="{opt}" {"selected" if values[field]==opt else ""}>{opt}</option>'
                for opt in dropdown_fields[field]
            ])
            input_field = f'<select name="{field}" {disabled_attr}>{options}</select>'
        else:
            input_field = f'<input type="text" name="{field}" value="{values[field]}" {disabled_attr}>'

        cell_html = f"<div class='form-cell'>{label}{input_field}</div>"
        row_html += cell_html
        count += 1

        if count % 3 == 0:
            form_html += f"<div class='form-row'>{row_html}</div>"
            row_html = ""

    if row_html:
        form_html += f"<div class='form-row'>{row_html}</div>"

    download_button = ""
    if session.get("filename"):
        download_button = f"""
        <form method="get" action="/download">
            <button type="submit">⬇️ دانلود فایل خروجی</button>
        </form>
        """

    return f"""
<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <title>فرم بیمه عمر</title>
    <style>
        body {{
            font-family: 'Tahoma';
            background-color: #6CABDD;
            padding: 20px;
            direction: rtl;
            text-align: right;
        }}
        .form-row {{
            display: flex;
            justify-content: space-between;
            margin-bottom: 15px;
        }}
        .form-cell {{
            flex: 1;
            margin-left: 10px;
        }}
        input, select {{
            width: 100%;
            padding: 5px;
            margin-top: 5px;
        }}
        label {{
            display: block;
            font-weight: bold;
        }}
        button {{
            margin-top: 20px;
            padding: 10px 20px;
            background-color: #cde6ff;
            border: none;
            font-size: 16px;
            cursor: pointer;
        }}
        .message {{
            margin-top: 20px;
            font-weight: bold;
        }}
        .top-bar {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 30px;
            background-color: #ffffff;
            padding: 15px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        .header-content {{
            display: flex;
            align-items: center;
            gap: 15px;
        }}
        .logo {{
            height: 60px;
            width: auto;
        }}
        .company-info {{
            text-align: right;
        }}
        .company-name {{
            font-size: 18px;
            font-weight: bold;
            color: #1a472a;
            margin: 0;
        }}
        .form-title {{
            font-size: 14px;
            color: #666;
            margin: 5px 0 0 0;
        }}
        .print-button {{
            background-color: #fff;
            border: 1px solid #333;
            padding: 5px 10px;
            cursor: pointer;
            font-size: 14px;
        }}
        .welcome-button {{
            float: left;
            margin-top: 30px;
            background-color: #f0f0f0;
            border: 1px solid #333;
            padding: 10px 20px;
            cursor: pointer;
            font-size: 14px;
        }}
        .samanyar-button {{
            float: left;
            margin-top: 30px;
            background-color: #fff;
            color: red;
            border: 1px solid red;
            padding: 10px 20px;
            cursor: pointer;
            font-size: 14px;
            margin-left: 10px;
        }}
    </style>
    <script>
        function updateInterest() {{
            const plan = document.querySelector('select[name="طرح"]');
            const interestMap = {{
                "پرریسک": "20.73%",
                "رشد1": "33.69%",
                "رشد": "22.69%",
                "کم_ریسک": "16.52%",
                "متوسط_ریسک": "17.18%"
            }};
            const interestField = document.querySelector('input[name="میانگین نرخ سود قطعی"]');
            if (interestField && !interestField.disabled) {{
                interestField.value = interestMap[plan.value] || "";
            }}
        }}
        window.onload = function() {{
            const planSelect = document.querySelector('select[name="طرح"]');
            if (planSelect) {{
                planSelect.addEventListener("change", updateInterest);
                updateInterest();
            }}
        }}
        function printForm() {{
            window.print();
        }}
        function openWelcome() {{
            const win = window.open("", "WelcomeWindow", "width=400,height=200");
            win.document.write("<h2 style='font-family:tahoma; text-align:center;'>خوش آمدید به برنامه کاربردی من</h2>");
        }}
    </script>
</head>
<body>
    <div class="top-bar">
        <div class="header-content">
            <img src="/attached_assets/saman_logo_1753971421258.png" class="logo" alt="بیمه سامان">
            <div class="company-info">
                <h3 class="company-name">بیمه سامان</h3>
                <p class="form-title">فرم بیمه عمر و تشکیل سرمایه</p>
            </div>
        </div>
        <button class="print-button" onclick="printForm()">🖨️ چاپ فرم</button>
    </div>
    <form method="post">
        {form_html}
        <button type="submit" name="action" value="ذخیره فرم">ذخیره فرم</button>
        <button type="submit" name="action" value="ویرایش فرم">ویرایش فرم</button>
    </form>
    {download_button}
    <div class="message">{message}</div>
    <button class="welcome-button" onclick="openWelcome()">👋 خوش آمدید</button>
    <button class="samanyar-button" onclick="window.location.href='/samanyar'">🔴 ورود به پنجره طرح سامانیار</button>
</body>
</html>
    """






from flask import Flask, request, session, send_file, render_template_string
from openpyxl import load_workbook
import os
import uuid


samanyar_mapping = {
    "سن بیمه گذار": "C6",
    "ضریب هزینه پزشکی ناشی از حادثه": "C28",
    "ضریب درآمد ازکارافتادگی": "C30",
    "منظور کردن پوشش ها در حق بیمه اولیه": "C32",
    "درصد اضافه نرخ پزشکی": "C34",
    "مبلغ پرداخت اضافه": "C36",
    "سن بیمه شده": "F6",
    "سرمایه فوت طرح سامانیار": "F8",
    "ضریب از کار افتادگی ناشی از حادثه": "F28",
    "دارای درآمد از کارافتادگی": "F30",
    "طرح امراض": "F32",
    "طبقه شغلی بیمه‌گذار": "F34",
    "شماره آخرین سال در پرداخت اضافه": "F36",
    "مدت بیمه نامه": "I6",
    "حق بیمه": "I8",
    "ضریب فوت حادثه": "I28",
    "دارای پوشش معافیت": "I30",
    "سرمایه امراض": "I32",
    "طبقه شغلی بیمه‌شده": "I34",
    "شماره اولین سال در پرداخت اضافه": "I36",
    "میانگین نرخ سود قطعی": "I38"
}

def generate_filename():
    return f"samanyar_output_{uuid.uuid4().hex}.xlsx"

@app.route("/samanyar", methods=["GET", "POST"])
def samanyar():
    values = {field: "" for field in samanyar_mapping.keys()}
    message = ""

    if request.method == "POST":
        action = request.form.get("action")
        values = {field: request.form.get(field, "") for field in samanyar_mapping.keys()}

        if action == "ذخیره فرم":
            try:
                filename = generate_filename()
                wb = load_workbook("attached_assets/__جدول__1753971258468.xlsx")
                sheet = wb["Info"]
                for field, cell in samanyar_mapping.items():
                    sheet[cell] = values[field]
                wb.save(filename)
                session["samanyar_locked"] = True
                session["samanyar_filename"] = filename
                message = "✅ فرم سامانیار ذخیره شد. حالا می‌تونی فایل رو دانلود کنی."
            except Exception as e:
                message = f"❌ خطا در ذخیره‌سازی: {e}"



        elif action == "ویرایش فرم":
            session["samanyar_locked"] = False
            session.pop("samanyar_filename", None)
            message = "✅ فرم قابل ویرایش شد."

    locked = session.get("samanyar_locked", False)
    return render_samanyar_form(values, locked, message)

@app.route("/download_samanyar")
def download_samanyar():
    filename = session.get("samanyar_filename")
    if filename and os.path.exists(filename):
        return send_file(filename, as_attachment=True)
    return "❌ فایل خروجی پیدا نشد."

def render_samanyar_form(values, locked, message):
    form_html = """
    <!DOCTYPE html>
    <html lang="fa">
    <head>
        <meta charset="UTF-8">
        <title>فرم سامانیار</title>
        <style>
            body {
                font-family: Vazir, sans-serif;
                direction: rtl;
                background: #6CABDD; /* آبی منچسترسیتی */
                padding: 20px;
            }
            .container {
                max-width: 800px;
                margin: auto;
                background: white;
                padding: 20px;
                border-radius: 12px;
                box-shadow: 0 0 15px rgba(0,0,0,0.2);
            }
            .top-bar {
                display: flex;
                justify-content: space-between;
                align-items: center;
                margin-bottom: 30px;
                background-color: #ffffff;
                padding: 15px;
                border-radius: 10px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            }
            .header-content {
                display: flex;
                align-items: center;
                gap: 15px;
            }
            .logo {
                height: 60px;
                width: auto;
            }
            .company-info {
                text-align: right;
            }
            .company-name {
                font-size: 18px;
                font-weight: bold;
                color: #003366;
                margin: 0;
            }
            .form-title {
                font-size: 14px;
                color: #666;
                margin: 5px 0 0 0;
            }
            h2 {
                text-align: center;
                color: #003366;
                margin-bottom: 25px;
            }
            .field {
                margin-bottom: 15px;
            }
            label {
                display: block;
                font-weight: bold;
                margin-bottom: 5px;
                color: #003366;
            }
            input[type="text"] {
                width: 100%;
                padding: 8px;
                border: 1px solid #ccc;
                border-radius: 6px;
                background-color: #f0f8ff;
            }
            .buttons {
                display: flex;
                justify-content: space-between;
                margin-top: 30px;
            }
            .buttons button {
                padding: 10px 20px;
                font-size: 16px;
                border: none;
                border-radius: 6px;
                cursor: pointer;
                font-weight: bold;
            }
            .save {
                background-color: #cde6ff;
                color: #003366;
            }
            .edit {
                background-color: #cde6ff;
                color: #003366;
            }
            .save:disabled, .edit:disabled {
                background-color: #e0e0e0;
                color: #888;
                cursor: not-allowed;
            }
            .message {
                margin-top: 20px;
                font-weight: bold;
                color: #003366;
                text-align: center;
            }
            .download {
                margin-top: 20px;
                text-align: center;
            }
            .download a {
                background-color: #cde6ff;
                color: #003366;
                padding: 10px 20px;
                border-radius: 6px;
                text-decoration: none;
                font-weight: bold;
                display: inline-block;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="top-bar">
                <div class="header-content">
                    <img src="/attached_assets/saman_logo_1753971421258.png" class="logo" alt="بیمه سامان">
                    <div class="company-info">
                        <h3 class="company-name">بیمه سامان</h3>
                        <p class="form-title">فرم اطلاعات طرح سامانیار</p>
                    </div>
                </div>
            </div>
            <h2>فرم اطلاعات طرح سامانیار</h2>
            <form method="POST">
                {% for field in fields %}
                    <div class="field">
                        <label>{{ field }}</label>
                        <input type="text" name="{{ field }}" value="{{ values[field] }}" {% if locked %}readonly{% endif %}>
                    </div>
                {% endfor %}
                <div class="buttons">
                    <button type="submit" name="action" value="ذخیره فرم" class="save" {% if locked %}disabled{% endif %}>ذخیره فرم</button>
                    <button type="submit" name="action" value="ویرایش فرم" class="edit" {% if not locked %}disabled{% endif %}>ویرایش فرم</button>
                </div>
            </form>
            <div class="message">{{ message }}</div>
            {% if locked %}
                <div class="download">
                    <a href="/download_samanyar">⬇️ دانلود فایل خروجی</a>
                </div>
            {% endif %}
        </div>
    </body>
    </html>
    """
    return render_template_string(form_html, fields=samanyar_mapping.keys(), values=values, locked=locked, message=message)





if __name__ == "__main__":
    app.run(debug=True)


